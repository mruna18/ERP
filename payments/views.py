from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from companies.models import Company
from invoice.models import *
from .models import *
from .serializers import *
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
import traceback
from django.http import HttpResponse
from openpyxl import Workbook
from invoice.models import Invoice
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from .utils import *
from django.utils.text import slugify
from django.conf import settings
from staff.permission import get_company_id, IsCompanyAdminOrAssigned, HasModulePermission

class CreatePaymentInView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Payment" 
    required_permission = "create"
    


    def post(self, request):
        data = request.data
        company_id = data.get('company')
        invoice_id = data.get('invoice')
        bank_id = data.get('bank_account')
        amount = float(data.get('amount', 0))

        # Validate company
        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return Response({"status": 500, "message": "Company not found"})

        # Validate invoice
        try:
            invoice = Invoice.objects.get(id=invoice_id, company=company)
        except Invoice.DoesNotExist:
            return Response({"status": 500, "message": "Invoice not found for this company"})

        # Bank account validation 
        bank = None
        if bank_id:
            try:
                bank = BankAccount.objects.get(id=bank_id, company=company, deleted=False)
            except BankAccount.DoesNotExist:
                return Response({"status": 500, "message": "Invalid or inactive bank account for this company"}, status=400)

        # Prevent overpayment
        if invoice.remaining_balance <= 0:
            return Response({
                "status": 500,
                "message": "Invoice is already fully paid. No further payment is required."
            }, status=400)

        if round(amount, 2) > round(invoice.remaining_balance, 2):
            return Response({
                "status": 500,
                "message": f"Payment amount exceeds the remaining balance of ₹{invoice.remaining_balance:.2f}."
            }, status=400)

        # Create PaymentIn record
        payment = PaymentIn.objects.create(
            company=company,
            invoice=invoice,
            amount=amount,
            bank_account=bank,
            note=data.get('note', '')
        )

       #? Handle Cash Ledger if payment_type is cash
        if invoice.payment_type and invoice.payment_type.name.lower() == 'cash':
            from payments.models import CashLedger, CashTransaction  # avoid circular import
            cash_ledger = CashLedger.objects.filter(company_name=company, deleted=False).first()
            if not cash_ledger:
                return Response({"status": 500, "message": "No active cash ledger found for this company"}, status=400)

            cash_ledger.current_balance += amount
            cash_ledger.save()

            CashTransaction.objects.create(
                ledger=cash_ledger,
                transaction_type='credit',
                amount=amount,
                description=f"Cash payment received for invoice #{invoice.invoice_number}",
                balance_after_transaction=cash_ledger.current_balance
            )

        #  Handle bank transaction if bank is provided
        if bank:
            bank.current_balance += amount
            bank.save()

            BankTransaction.objects.create(
                bank_account=bank,
                transaction_type='credit',
                amount=amount,
                related_invoice=invoice,
                description=data.get('note', f"Payment In for invoice #{invoice.invoice_number}"),
                balance_after_transaction=bank.current_balance
            )

        # Update invoice
        invoice.amount_paid += amount
        invoice.remaining_balance = max(invoice.total - invoice.amount_paid, 0)

        if invoice.amount_paid == 0:
            invoice.payment_status_id = 1  # Unpaid
        elif invoice.amount_paid < invoice.total:
            invoice.payment_status_id = 2  # Partially Paid
        else:
            invoice.payment_status_id = 3  # Paid


        # Overpayment detection
        overpaid_amount = round(invoice.amount_paid - invoice.total, 2)
        overpayment_warning = None

        if overpaid_amount > 0:
            overpayment_warning = f"You have overpaid ₹{overpaid_amount:.2f}. Please verify."

        invoice.save()

        return Response({
            "status": 200,
            "message": "Payment In recorded and invoice updated successfully",
            "data": PaymentInSerializer(payment).data,
            "warning": overpayment_warning 
        })


class CreatePaymentOutView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Payment" 
    required_permission = "create"
   

    def post(self, request):
        data = request.data
        company_id = data.get("company")
        invoice_id = data.get("invoice")
        amount = float(data.get("amount", 0))
        bank_account_id = data.get("bank_account")
        note = data.get("note", "")
        payment_type_id = data.get("payment_type")  # new

        # Validate company
        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return Response({"status": 500, "message": "Company not found."})

        # Validate invoice (optional)
        invoice = None
        if invoice_id:
            try:
                invoice = Invoice.objects.get(id=invoice_id, company=company)
            except Invoice.DoesNotExist:
                return Response({"status": 500, "message": "Invoice not found for this company."})

        # Validate payment type
        try:
            payment_type = PaymentType.objects.get(id=payment_type_id)
        except PaymentType.DoesNotExist:
            return Response({"status": 500, "message": "Invalid payment type."}, status=400)

      
       #? Handle Cash Payments
      
        if payment_type.name.lower() == 'cash':
            try:
                ledger = CashLedger.objects.get(company_name=company, deleted=False)
            except CashLedger.DoesNotExist:
                return Response({"status": 500, "message": "Cash ledger not found for this company."})

            if ledger.current_balance < amount:
                return Response({"status": 500, "message": "Insufficient cash balance."}, status=400)

            # Create PaymentOut (cash)
            payment = PaymentOut.objects.create(
                company=company,
                invoice=invoice,
                amount=amount,
                bank_account=None,
                note=note
            )

            # Update cash ledger
            ledger.current_balance -= amount
            ledger.save()

            # Create CashTransaction
            CashTransaction.objects.create(
                ledger=ledger,
                transaction_type='debit',
                amount=amount,
                description=note or f"Cash payment for invoice #{invoice.invoice_number}" if invoice else "Cash payment",
                balance_after_transaction=ledger.current_balance
            )

      
       #? Handle Bank Payments
      
        else:
            try:
                bank_account = BankAccount.objects.get(id=bank_account_id, company=company, deleted=False)
            except BankAccount.DoesNotExist:
                return Response({"status": 500, "message": "Bank account not found for this company."})

            if bank_account.current_balance < amount:
                return Response({"status": 500, "message": "Insufficient bank balance."}, status=400)

            # Create PaymentOut (bank)
            payment = PaymentOut.objects.create(
                company=company,
                invoice=invoice,
                amount=amount,
                bank_account=bank_account,
                note=note
            )

            # Update bank balance
            bank_account.current_balance -= amount
            bank_account.save()

            # Create BankTransaction
            BankTransaction.objects.create(
                bank_account=bank_account,
                transaction_type='debit',
                amount=amount,
                related_invoice=invoice,
                description=note or f"Payment Out (Invoice #{invoice.invoice_number})" if invoice else "Payment Out",
                balance_after_transaction=bank_account.current_balance
            )

        return Response({
            "status": 200,
            "message": "PaymentOut created successfully.",
            "data": {
                "id": payment.id,
                "amount": payment.amount,
                "invoice": invoice.id if invoice else None,
                "bank_account": payment.bank_account.id if payment.bank_account else None,
                "note": payment.note
            }
        })


class ListPaymentsView(APIView):
    """List PaymentIn and PaymentOut for a company (POST with company id)."""
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Payment"
    required_permission = "get_using_post"

    def post(self, request):
        company_id = get_company_id(request, self)
        if not company_id:
            return Response({"detail": "Company ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        payment_ins = (
            PaymentIn.objects
            .filter(company_id=company_id)
            .select_related("invoice", "invoice__party", "bank_account")
            .order_by("-payment_date", "-id")
        )
        payment_outs = (
            PaymentOut.objects
            .filter(company_id=company_id)
            .select_related("invoice", "invoice__party", "bank_account")
            .order_by("-payment_date", "-id")
        )

        def build_in(p):
            return {
                "id": p.id,
                "type": "in",
                "amount": p.amount,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "note": p.note or "",
                "invoice_id": p.invoice_id,
                "invoice_number": p.invoice.invoice_number if p.invoice else None,
                "party_name": p.invoice.party.name if p.invoice and p.invoice.party else None,
                "bank_name": p.bank_account.bank_name if p.bank_account else None,
            }

        def build_out(p):
            return {
                "id": p.id,
                "type": "out",
                "amount": p.amount,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "note": p.note or "",
                "invoice_id": p.invoice_id,
                "invoice_number": p.invoice.invoice_number if p.invoice else None,
                "party_name": p.invoice.party.name if p.invoice and p.invoice.party else None,
                "bank_name": p.bank_account.bank_name if p.bank_account else "Cash",
            }

        return Response({
            "payment_ins": [build_in(p) for p in payment_ins],
            "payment_outs": [build_out(p) for p in payment_outs],
        })


#! -- cash ledger --
class CreateCashLedgerView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash Ledger" 
    required_permission = "create"
  


    def post(self, request):
        try:
            data = request.data

            company_id = data.get("company_name")
            ledger_name = data.get("ledger_name")
            opening_balance = float(data.get("opening_balance", 0))
            as_on = data.get("as_on")

            if not company_id or not ledger_name:
                return Response({"msg": "company_name and ledger_name are required", "status": 400})

            company = Company.objects.filter(id=company_id).first()
            if not company:
                return Response({"msg": "Company not found", "status": 404})

            if CashLedger.objects.filter(ledger_name=ledger_name, company_name=company, deleted=False).exists():
                return Response({"msg": "Cash Ledger already exists", "status": 400})

            ledger = CashLedger.objects.create(
                ledger_name=ledger_name,
                company_name=company,
                opening_balance=opening_balance,
                current_balance=opening_balance,
                as_on=as_on
            )

            return Response({
                "msg": "Cash Ledger Created Successfully",
                "data": {
                    "id": ledger.id,
                    "ledger_name": ledger.ledger_name,
                    "company_name": ledger.company_name.id,
                    "opening_balance": ledger.opening_balance,
                    "current_balance": ledger.current_balance,
                    "as_on": ledger.as_on
                },
                "status": 201
            })

        except Exception as e:
            import traceback
            return Response({"msg": traceback.format_exc(), "status": 500})
        

# cash transaction 
class CreateCashTransactionView(APIView):
    pepermission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash" 
    required_permission = "create"
    

    def post(self, request):
        try:
            data = request.data
            ledger_id = data.get("ledger")
            transaction_type = data.get("transaction_type")
            amount = float(data.get("amount", 0))
            description = data.get("description", "")

            if not ledger_id or not transaction_type or amount <= 0:
                return Response({"status": 500, "message": "Invalid input data."}, status=400)

            try:
                ledger = CashLedger.objects.get(id=ledger_id, deleted=False)
            except CashLedger.DoesNotExist:
                return Response({"status": 500, "message": "Cash ledger not found."})

            # Compute new balance
            if transaction_type == 'credit':
                new_balance = ledger.current_balance + amount
            elif transaction_type == 'debit':
                if amount > ledger.current_balance:
                    return Response({
                        "status": 500,
                        "message": "Insufficient cash balance for debit."
                    }, status=400)
                new_balance = ledger.current_balance - amount
            else:
                return Response({"status": 500, "message": "Invalid transaction type."}, status=400)

            with transaction.atomic():
                # Create transaction
                txn = CashTransaction.objects.create(
                    ledger=ledger,
                    transaction_type=transaction_type,
                    amount=amount,
                    description=description,
                    balance_after_transaction=new_balance
                )

                # Update ledger balance
                ledger.current_balance = new_balance
                ledger.save()

            return Response({
                "status": 200,
                "message": "Cash transaction created successfully.",
                "data": CashTransactionSerializer(txn).data
            })

        except Exception:
            return Response({
                "status": 500,
                "message": traceback.format_exc()
            })
        

class UpdateCashLedgerView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash Ledger" 
    required_permission = "edit"
    

    def put(self, request, pk):
        try:
            ledger = CashLedger.objects.get(pk=pk, deleted=False)
        except CashLedger.DoesNotExist:
            return Response({"msg": "Cash ledger not found", "status": 404})

        serializer = CashLedgerSerializer(ledger, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"msg": "Cash ledger updated successfully", "data": serializer.data, "status": 200})
        else:
            return Response({"msg": serializer.errors, "status": 400}, status=400)
        


class DeleteCashLedgerView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash Ledger" 
    required_permission = "delete"
 

    def delete(self, request, pk):
        try:
            ledger = CashLedger.objects.get(pk=pk, deleted=False)
        except CashLedger.DoesNotExist:
            return Response({"msg": "Cash ledger not found", "status": 404})

        ledger.deleted = True
        ledger.save()
        return Response({"msg": "Cash ledger soft-deleted successfully", "status": 200})


# list of all ledger
class ListCashLedgersView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash Ledger" 
    required_permission = "get_using_post"
    

    def get(self, request, company_id):
        ledgers = CashLedger.objects.filter(company_name_id=company_id, deleted=False)
        serializer = CashLedgerSerializer(ledgers, many=True)
        return Response({"msg": "Success", "data": serializer.data, "status": 200})
    
# get ledger by id
class GetCashLedgerByIdView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Cash Ledger" 
    required_permission = "view_specific"
    

    def get(self, request, pk):
        try:
            ledger = CashLedger.objects.get(pk=pk, deleted=False)
        except CashLedger.DoesNotExist:
            return Response({"msg": "Cash ledger not found", "status": 404})

        serializer = CashLedgerSerializer(ledger)
        return Response({"msg": "Success", "data": serializer.data, "status": 200})
    
#! -- bank to bank

class BankToBankTransferView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Bank Transfer" 
    required_permission = "create"


    def post(self, request):
        data = request.data
        company_id = data.get("company")
        from_account_id = data.get("from_account")
        to_account_id = data.get("to_account")
        amount = float(data.get("amount", 0))
        note = data.get("note", "")

        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return Response({"status": 500, "message": "Company not found."})

        if from_account_id == to_account_id:
            return Response({"status": 500, "message": "From and To accounts cannot be the same."}, status=400)

        try:
            from_account = BankAccount.objects.get(id=from_account_id, company=company, deleted=False)
            to_account = BankAccount.objects.get(id=to_account_id, company=company, deleted=False)
        except BankAccount.DoesNotExist:
            return Response({"status": 500, "message": "Invalid bank accounts for this company."}, status=400)

        if from_account.current_balance < amount:
            return Response({"status": 500, "message": "Insufficient balance in source account."}, status=400)

        with transaction.atomic():
            # Transfer record
            transfer = BankToBankTransfer.objects.create(
                company=company,
                from_account=from_account,
                to_account=to_account,
                amount=amount,
                note=note
            )

            # Update balances
            from_account.current_balance -= amount
            to_account.current_balance += amount
            from_account.save()
            to_account.save()

            # Log debit transaction in source
            BankTransaction.objects.create(
                bank_account=from_account,
                transaction_type='debit',
                amount=amount,
                # description=f"Transfer to {to_account.account_name}. {note}",
                description=f"Transfer to {to_account.bank_name} (A/C {to_account.account_no}). {note}",
                balance_after_transaction=from_account.current_balance
            )

            # Log credit transaction in destination
            BankTransaction.objects.create(
                bank_account=to_account,
                transaction_type='credit',
                amount=amount,
                # description=f"Transfer from {from_account.account_name}. {note}",
                description=f"Transfer from {to_account.bank_name} (A/C {to_account.account_no}). {note}",
                balance_after_transaction=to_account.current_balance
            )

        return Response({
            "status": 200,
            "message": "Transfer successful.",
            "transfer_id": transfer.id
        })
    

# List all bank transfers for a specific company
class ListBankTransfersView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Bank Transfer" 
    required_permission = "get_using_post"
   

    def get(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return Response({"status": 404, "message": "Company not found."}, status=404)

        transfers = BankToBankTransfer.objects.filter(company=company, deleted=False).order_by('-created_at')
        serializer = BankTransferSerializer(transfers, many=True)
        return Response({
            "status": 200,
            "message": f"Bank transfers for company '{company.name}' retrieved successfully.",
            "data": serializer.data
        }, status=200)


# Get specific transfer by ID
class GetBankTransferView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Bank Transfer" 
    required_permission = "view_specific"
    

    def get(self, request, pk):
        transfer = BankToBankTransfer.objects.filter(id=pk, deleted=False).first()
        if not transfer:
            return Response({"status": 404, "message": "Bank transfer not found."}, status=404)
        serializer = BankTransferSerializer(transfer)
        return Response({
            "status": 200,
            "message": "Bank transfer details retrieved successfully.",
            "data": serializer.data
        }, status=200)


#  Update a specific transfer
class UpdateBankToBankTransferView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Bank Transfer" 
    required_permission = "edit"
    

    def put(self, request, transfer_id):
        data = request.data
        company_id = data.get("company")
        new_from_id = data.get("from_account")
        new_to_id = data.get("to_account")
        new_amount = float(data.get("amount", 0))
        note = data.get("note", "")

        try:
            transfer = BankToBankTransfer.objects.get(id=transfer_id, company_id=company_id, deleted=False)
        except BankToBankTransfer.DoesNotExist:
            return Response({"status": 404, "message": "Transfer not found."})

        if new_from_id == new_to_id:
            return Response({"status": 500, "message": "From and To accounts cannot be the same."})

        try:
            new_from = BankAccount.objects.get(id=new_from_id, company_id=company_id, deleted=False)
            new_to = BankAccount.objects.get(id=new_to_id, company_id=company_id, deleted=False)
        except BankAccount.DoesNotExist:
            return Response({"status": 500, "message": "Invalid bank accounts."})

        old_from = transfer.from_account
        old_to = transfer.to_account
        old_amount = transfer.amount

        with transaction.atomic():
            # Revert old transfer
            old_from.current_balance += old_amount
            old_to.current_balance -= old_amount
            old_from.save()
            old_to.save()

            # Check if new source account has sufficient funds
            if new_from.current_balance < new_amount:
                return Response({"status": 500, "message": "Insufficient funds in new source account."})

            # Apply new transfer
            new_from.current_balance -= new_amount
            new_to.current_balance += new_amount
            new_from.save()
            new_to.save()

            # Update transfer record
            transfer.from_account = new_from
            transfer.to_account = new_to
            transfer.amount = new_amount
            transfer.note = note
            transfer.updated_at = timezone.now()
            transfer.save()

            # Create transactions
            BankTransaction.objects.create(
                bank_account=new_from,
                transaction_type='debit',
                amount=new_amount,
                related_invoice=None,
                description=f"Updated transfer to {new_to.bank_name} (A/C {new_to.account_no}). {note}",
                balance_after_transaction=new_from.current_balance
            )

            BankTransaction.objects.create(
                bank_account=new_to,
                transaction_type='credit',
                amount=new_amount,
                related_invoice=None,
                description=f"Updated transfer from {new_from.bank_name} (A/C {new_from.account_no}). {note}",
                balance_after_transaction=new_to.current_balance
            )

        return Response({
            "status": 200,
            "message": "Bank-to-bank transfer updated successfully.",
            "transfer_id": transfer.id
        })

# Soft delete a specific transfer
class DeleteBankToBankTransferView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Bank Transfer" 
    required_permission = "delete"
    

    def delete(self, request, transfer_id):
        try:
            transfer = BankToBankTransfer.objects.get(id=transfer_id, deleted=False)
        except BankToBankTransfer.DoesNotExist:
            return Response({"status": 404, "message": "Transfer not found or already deleted."})

        from_account = transfer.from_account
        to_account = transfer.to_account
        amount = transfer.amount

        with transaction.atomic():
            # Revert balances
            from_account.current_balance += amount
            to_account.current_balance -= amount
            from_account.save()
            to_account.save()

            # Soft delete the transfer
            transfer.deleted = True
            transfer.save()

        return Response({
            "status": 200,
            "message": f"Transfer #{transfer.id} deleted and balances reverted successfully."
        })


#! -- excel --

class SalesReportExportExcelView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Reports"
    required_permission = "view"
    

    def post(self, request):
        company_id = request.data.get("company")
        payment_status = request.data.get("payment_status")

        if not all([company_id, payment_status]):
            return Response({
                "status": 500,
                "message": "company and payment_status are required."
            }, status=400)

        # Default invoice_type = 1 (Sales)
        invoices = Invoice.objects.filter(
            company_id=company_id,
            invoice_type_id=1,
            payment_status_id=payment_status,
            is_deleted=False
        )

        if not invoices.exists():
            return Response({
                "status": 404,
                "message": "No data found for the selected filters. Nothing to export."
            }, status=404)

        # === Create Workbook ===
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # === Header Row 1 (Title) ===
        ws.merge_cells("A1:I1")
        ws["A1"].value = "Sales Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # === Header Row 2 (Company + Date) ===
        company_name = invoices.first().company.name
        today = datetime.now().strftime("%d %B %Y")
        ws.merge_cells("A2:I2")
        ws["A2"].value = f"{company_name} - {today}"
        ws["A2"].font = Font(bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # === Table Headers (Row 3) ===
        headers = [
            "Invoice Number", "Date", "Party", "Subtotal", "Tax", "Discount",
            "Total", "Amount Paid", "Payment Status"
        ]
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # === Data Rows ===
        for row_num, invoice in enumerate(invoices, start=4):
            row_data = [
                invoice.invoice_number,
                invoice.created_at.strftime("%Y-%m-%d"),
                invoice.party.name if invoice.party else "",
                float(invoice.subtotal),
                float(invoice.tax_amount),
                float(invoice.discount_amount),
                float(invoice.total),
                float(invoice.amount_paid),
                invoice.payment_status.label if invoice.payment_status else ""
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border

        # === Auto Width Columns ===
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # === Unique Filename Handling ===
        year = datetime.now().year
        base_name = f"{slugify(company_name)}-SalesReport-{year}"
        filename = f"{base_name}.xlsx"
        media_dir = settings.MEDIA_ROOT
        filepath = os.path.join(media_dir, filename)

        counter = 1
        while os.path.exists(filepath):
            filename = f"{base_name}({counter}).xlsx"
            filepath = os.path.join(media_dir, filename)
            counter += 1

        wb.save(filepath)
        download_url = request.build_absolute_uri(f"{settings.MEDIA_URL}{filename}")

        return Response({
            "msg": "Exported in Excel successfully",
            "file_url": download_url,
            "status": 200
        })


#! -- purchase excel --
class PurchaseReportExportExcelView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Reports"
    required_permission = "view"
 

    def post(self, request):
        company_id = request.data.get("company")
        payment_status = request.data.get("payment_status")
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

        if not company_id or not payment_status:
            return Response({"status": 400, "message": "company and payment_status are required."})

        # Fetch only purchase-type invoices (assuming id=2 OR use code='purchase')
        invoices = Invoice.objects.filter(
            company_id=company_id,
            invoice_type__code__iexact="purchase",
            payment_status_id=payment_status,
            is_deleted=False
        )

        if not invoices.exists():
            return Response({
                "status": 404,
                "message": "No data found for the selected filters. Nothing to export."
            }, status=404)


        if start_date and end_date:
            invoices = invoices.filter(created_at__date__range=[start_date, end_date])

        if not invoices.exists():
            return Response({"status": 404, "message": "No matching purchase invoices found."})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Report"

        # Title
        title = "Purchase Report"
        ws.merge_cells("A1:I1")
        ws["A1"].value = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Company info
        company_name = invoices.first().company.name
        today = datetime.now().strftime("%d %B %Y")
        ws.merge_cells("A2:I2")
        ws["A2"].value = f"{company_name} - {today}"
        ws["A2"].font = Font(bold=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "Invoice Number", "Date", "Supplier", "Subtotal", "Tax", "Discount",
            "Total", "Amount Paid", "Payment Status"
        ]
        header_row = 3
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for row_num, invoice in enumerate(invoices, start=header_row + 1):
            row_data = [
                invoice.invoice_number,
                invoice.created_at.strftime("%Y-%m-%d"),
                invoice.party.name if invoice.party else "",
                float(invoice.subtotal),
                float(invoice.tax_amount),
                float(invoice.discount_amount),
                float(invoice.total),
                float(invoice.amount_paid),
                invoice.payment_status.label if invoice.payment_status else ""
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border

        # Auto column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # Save & respond with download URL
        year = datetime.now().year
        safe_name = company_name.replace(" ", "_")
        base_filename = f"{safe_name}-PurchaseReport-{year}.xlsx"
        # filepath = os.path.join(settings.MEDIA_ROOT, filename)

        filepath = get_unique_filepath(settings.MEDIA_ROOT, base_filename)

        # Save workbook
        wb.save(filepath)
                
        download_url = request.build_absolute_uri(f"{settings.MEDIA_URL}{filepath}")

        return Response({
            "msg": "Exported purchase report to Excel successfully.",
            "file_url": download_url,
            "status": 200
        })
    
#! -- book transaction excel --
class BankTransactionReportExportExcelView(APIView):
    permission_classes = [IsAuthenticated, IsCompanyAdminOrAssigned, HasModulePermission]
    required_module = "Reports"
    required_permission = "view"
    

    def post(self, request):
        company_id = request.data.get("company")
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

        if not company_id:
            return Response({"status": 400, "message": "Company ID is required."})

        transactions = BankTransaction.objects.filter(
            bank_account__company_id=company_id
        )

        if start_date and end_date:
            transactions = transactions.filter(
                created_at__date__range=[start_date, end_date]
            )

        if not transactions.exists():
            return Response({"status": 404, "message": "No bank transactions found for given filters."})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bank Transactions"

        ws.merge_cells("A1:G1")
        ws["A1"] = "Bank Transaction Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        company_name = transactions.first().bank_account.company.name
        today = datetime.now().strftime("%d %B %Y")
        ws.merge_cells("A2:G2")
        ws["A2"] = f"{company_name} - {today}"
        ws["A2"].font = Font(bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "Bank Account", "Type", "Amount", "Balance After Transaction",
            "Description", "Invoice", "Date"
        ]

        header_row = 3
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_num, tx in enumerate(transactions, start=header_row + 1):
            row_data = [
                str(tx.bank_account),
                tx.transaction_type.capitalize(),
                tx.amount,
                tx.balance_after_transaction,
                tx.description or "",
                tx.related_invoice.invoice_number if tx.related_invoice else "",
                tx.created_at.strftime("%Y-%m-%d %H:%M")
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

        year = datetime.now().year
        safe_name = company_name.replace(" ", "_")
        filename = f"{safe_name}-BankTransactions-{year}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        file_path = os.path.join(settings.MEDIA_ROOT, filename)
        wb.save(file_path)
        download_url = request.build_absolute_uri(f"{settings.MEDIA_URL}{filename}")

        return Response({
            "msg": "Bank transaction report exported successfully",
            "file_url": download_url,
            "status": 200
        })
